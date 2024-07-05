from django.shortcuts import render,get_object_or_404
from .forms import TraderForm,FarmerForm,SupplierForm,ItemForm,RegisterForm,UpdateTraderForm,UpdateFarmerForm
from django.http import HttpResponse, HttpResponseRedirect,JsonResponse
from . import models
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth import logout,login,authenticate
from main.decorators import check_login_and_redirect,check_user_client_redirect
from django.contrib.auth.hashers import make_password
from django.urls import reverse
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from datetime import datetime
from django.db.models import F, ExpressionWrapper, fields,Sum

def get_path(request):
    return request.path

# Default Views
@check_login_and_redirect
def login_view(request):
    context = {}
    message = ""
    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")

        instance_of_user = models.CustomUser.objects.filter(username=username)

        if instance_of_user.exists():
            auth = authenticate(request, username=username,password=password)

            if auth is not None:
                login(request,auth)

                if instance_of_user[0].is_superuser and instance_of_user[0].is_staff:
                    return HttpResponseRedirect("/ams/dashboard/")
                else:
                    return HttpResponseRedirect("/")
            else:
                message = "invalid_account"
        else:
            message = "invalid_username"

    context['message'] = message
    return render(request, "login.html",context)

@login_required(login_url="/ams/")
def export_stock_info(request):

    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    columns = [
        ("A", "PRODUCT ID"),
        ("B", "PRODUCT"),
        ("C", "PRODUCT/ITEM STATUS"),
        ("D", "PRODUCT SUPPLIER"),
        ("E", "STOCK"),
        ("F", "MEASUREMENT"),
        ("G", "WEIGHT"),
        ("H", "DATE ADDED")
    ]

    for col, header in columns:
        cell = worksheet["{}1".format(col)]
        cell.value = header
        cell.alignment = Alignment(horizontal='center')
        column_letter = get_column_letter(cell.column)
        max_length = max(len(str(cell.value) if cell.value is not None else 0) for cell in worksheet[column_letter])
        adjusted_width = (max_length + 4)
        worksheet.column_dimensions[column_letter].width = adjusted_width

    row_num = 2

    for product in models.Item.objects.all():
        worksheet.cell(row=row_num, column=1, value=product.item_id)
        worksheet.cell(row=row_num, column=2, value=product.item_name)
        worksheet.cell(row=row_num, column=3, value=product.item_status)
        worksheet.cell(row=row_num, column=4, value=product.item_supplier.supplier_name)
        worksheet.cell(row=row_num, column=5, value=product.item_quantity)
        worksheet.cell(row=row_num, column=6, value=product.item_measurement)
        worksheet.cell(row=row_num, column=7, value=product.item_kilo_per_measurement)
        worksheet.cell(row=row_num, column=8, value=product.item_date_added.strftime('%d/%m/%Y'))
        row_num += 1
    
    filename_str = f"stock_inventory_{datetime.now()}.xlsx"

    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename="{filename_str}"'
    
    workbook.save(response)

    return response

@check_login_and_redirect
def register_view(request):
    context = {}
    message = ""
    if request.method == "POST":

        register_form = RegisterForm(request.POST)

        if register_form.is_valid():
            if register_form.cleaned_data['password1'] != register_form.cleaned_data['password2']:
                message = "password_unmatched"
            else:

                if models.CustomUser.objects.filter(username=register_form.cleaned_data['username']).exists():
                    message = "existed"
                else:
                    user = models.CustomUser(
                        username=register_form.cleaned_data['username'],
                        first_name = register_form.cleaned_data['first_name'],
                        last_name = register_form.cleaned_data['last_name'],
                        email = register_form.cleaned_data['email'],
                        user_type= register_form.cleaned_data['user_type'],
                        contact_number = register_form.cleaned_data['contact_number'],
                        barangay= register_form.cleaned_data['barangay'],
                        address = register_form.cleaned_data['address'],
                        postal_code = register_form.cleaned_data['postal_code'],
                        password = make_password(register_form.cleaned_data['password1']),
                    )

                    user.save()

                    message = "success"
    print(message)    
    context['register_path'] = get_path(request=request)
    context['form'] = RegisterForm()
    context['message'] = message

    return render(request,"register.html",context=context)

def logout_account(request):
    logout(request=request)
    return HttpResponseRedirect("/ams/")

@login_required(login_url="/ams/")
@check_user_client_redirect
def dashboard_view(request):
    context = {}
    context['order_count']   = models.Transaction.objects.all().count()
    context['total_stock']    = models.Item.objects.aggregate(total_stock=Sum('item_quantity'))['total_stock']
    context['stock_sold']    = models.Cart.objects.aggregate(stock_sold=Sum('order_quantity'))['stock_sold']
    context['farmer_count']  = models.CustomUser.objects.filter(user_type="FARMER").count()
    context['trader_count']  = models.CustomUser.objects.filter(user_type="TRADER").count()
    context['product_count'] = models.Item.objects.all().count()
    context['supplier_count'] = models.Supplier.objects.all().count()
    context['orders'] = models.Transaction.objects.all()
    context['dashboard_path'] = get_path(request=request)
    return render(request, "dashboard.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def items_view(request):
    context = {}

    if request.method == "POST":
        item_instance = models.Item.objects.get(item_id=request.POST.get("item_id"))
        
        return JsonResponse({
            'item_id': item_instance.item_id,
            'item_name': item_instance.item_name,
            'item_desc': item_instance.item_desc,
            'item_status': item_instance.item_status,
            'item_quantity': item_instance.item_quantity,
            'item_date_added': item_instance.item_date_added,
        })
    
    context['items_path'] = get_path(request=request)
    context['item_lists'] = models.Item.objects.all()
    return render(request, "items.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def supplier_view(request):
    context = {}

    if request.method == "POST":
        supplier_instance = models.Supplier.objects.get(supplier_id=request.POST.get("supplier_id"))
        
        return JsonResponse({
            'supplier_id': supplier_instance.supplier_id,
            'supplier_name': supplier_instance.supplier_name,
            'address': supplier_instance.supplier_address,
            'postal_code': supplier_instance.supplier_postal_code,
            'barangay': supplier_instance.supplier_barangay,
            'contact_number': supplier_instance.supplier_mobile_number,
        })
    
    context['supplier_path'] = get_path(request=request)
    context['supplier_lists'] = models.Supplier.objects.all()
    return render(request, "supplier.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def farmer_view(request):
    context = {}

    if request.method == "POST":
        farmer_instance = models.CustomUser.objects.get(id=request.POST.get("farmer_id"))
        
        return JsonResponse({
            'farmer_id': farmer_instance.id,
            'first_name': farmer_instance.first_name,
            'last_name': farmer_instance.last_name,
            'username': farmer_instance.username,
            'user_type': farmer_instance.user_type,
            'contact_number': farmer_instance.contact_number,
            'barangay': farmer_instance.barangay,
            'postal_code': farmer_instance.postal_code,
            'address': farmer_instance.address,
        })
    
    context['farmer_path'] = get_path(request=request)
    context['farmer_lists'] = models.CustomUser.objects.filter(user_type="FARMER")

    return render(request, "farmer.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def trader_view(request):
    context = {}

    if request.method == "POST":
        trader_instance = models.CustomUser.objects.get(id=request.POST.get("trader_id"))
        
        return JsonResponse({
            'trader_id': trader_instance.id,
            'first_name': trader_instance.first_name,
            'last_name': trader_instance.last_name,
            'username': trader_instance.username,
            'user_type': trader_instance.user_type,
            'contact_number': trader_instance.contact_number,
            'barangay': trader_instance.barangay,
            'postal_code': trader_instance.postal_code,
            'address': trader_instance.address,
        })

    context['trader_path'] = get_path(request=request)
    context['trader_lists'] = models.CustomUser.objects.filter(user_type="TRADER")
    return render(request,"trader.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def transaction_view(request):
    context = {}


    if request.method == "POST":
        order_instance = models.Transaction.objects.get(id=request.POST.get("transaction_id"))
        
        return JsonResponse({
            'order_id': order_instance.id,
            'order_by__first_name': order_instance.user_transaction.first_name,
            'order_by__last_name': order_instance.user_transaction.last_name,
            'order_by__postal_code': order_instance.user_transaction.postal_code,
            'order_by__barangay': order_instance.user_transaction.barangay,
            'order_by__contact_number': order_instance.user_transaction.contact_number,
            'quantity': order_instance.quantity_per_product,
            'product_name': order_instance.order_product.item_name,
            'date_ordered': order_instance.date_transaction,
        })
    
    context['transactions'] = models.Transaction.objects.all()
    context['order_path'] = get_path(request=request)
    return render(request,"transaction.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def cart_view(request):
    context = {}
   
    context['cart_lists'] = models.Cart.objects.filter(is_checkedOut=False)
    context['cart_path'] = get_path(request=request)
    return render(request,"cart.html",context=context)

# Add Views
@login_required(login_url="/ams/")
@check_user_client_redirect
def add_supplier_view(request):
    context = {}

    if request.method == "POST":
        add_supplier_form = SupplierForm(request.POST)

        if add_supplier_form.is_valid():
            add_supplier_form.save()
            messages.success(request,"You have successfully added new supplier.")
            return HttpResponseRedirect("/ams/supplier/")
            
    context['add_supplier_path'] = get_path(request=request)
    context['form'] = SupplierForm()
    return render(request, "add_supplier.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def add_trader_view(request):
    context = {}

    if request.method == "POST":
        add_trader_form = TraderForm(request.POST)

        if add_trader_form.is_valid():
            trader = add_trader_form.save(commit=False)
            trader.set_password(add_trader_form.cleaned_data['password'])
            trader.save()
            messages.success(request,"You have successfully added new trader.")
            return HttpResponseRedirect("/ams/trader/")
        
    context['add_trader_path'] = get_path(request=request)
    context['form'] = TraderForm(initial={'user_type':'TRADER'})

    return render(request,"add_trader.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def add_farmer_view(request):
    context = {}

    if request.method == "POST":
        add_farmer_form = FarmerForm(request.POST)
        
        if add_farmer_form.is_valid():
            farmer = add_farmer_form.save()
            farmer.set_password(add_farmer_form.cleaned_data['password'])
            farmer.save()
            messages.success(request,"You have successfully added new farmer information.")
            return HttpResponseRedirect("/ams/farmer/")

    context['add_farmer_path'] = get_path(request=request)
    context['form'] = FarmerForm()
    return render(request, "add_farmer.html",context=context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def add_item_view(request):
    context = {}

    if request.method == "POST":
        add_item_form = ItemForm(request.POST,request.FILES)

        if add_item_form.is_valid():
            add_item_form.save()
            messages.success(request,"You have successfully added new item information.")
            return HttpResponseRedirect("/ams/items/")

    context['add_item_path'] = get_path(request=request)
    context['form'] = ItemForm()
    return render(request, "add_item.html",context=context)


# Edit View
@login_required(login_url="/ams/")
@check_user_client_redirect
def edit_supplier_view(request,pk):
    context = {}

    instance = get_object_or_404(models.Supplier,supplier_id=pk)

    if request.method == "POST":
        update_supplier_form = SupplierForm(request.POST,instance=instance)
        
        if update_supplier_form.is_valid():
            update_supplier_form.save()
            messages.success(request,"You have successfully updated supplier information.")
            return HttpResponseRedirect("/ams/trader/")
    else:
        context['form'] = SupplierForm(instance=instance)

    context['supplier_id'] = pk 
    return render(request,"edit_supplier.html",context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def edit_trader_view(request,pk):
    context = {}

    instance = get_object_or_404(models.CustomUser,id=pk)


    if request.method == "POST":
        update_trader_form = UpdateTraderForm(request.POST,instance=instance)
        
        if update_trader_form.is_valid():
            update_trader_form.save()
            messages.success(request,"You have successfully updated trader information.")
            return HttpResponseRedirect("/ams/trader/")

    else:
        context['form'] = UpdateTraderForm(instance=instance)

    context['trader_id'] = pk
    return render(request,"edit_trader.html",context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def edit_farmer_view(request,pk):
    context = {}

    instance = get_object_or_404(models.CustomUser,id=pk)

    if request.method == "POST":
        data = request.POST

        update_farmer_form = UpdateFarmerForm(data,instance=instance)
        if update_farmer_form.is_valid():
            update_farmer_form.save()
            messages.success(request,"You have successfully updated farmer information.")
            return HttpResponseRedirect("/ams/farmer/")
    else:
        context['form'] = UpdateFarmerForm(instance=instance)

    context['farmer_id'] = pk
    return render(request,"edit_farmer.html",context)

@login_required(login_url="/ams/")
@check_user_client_redirect
def edit_item_view(request,pk):
    context = {}


    instance = get_object_or_404(models.Item,item_id=pk)

    if request.method == "POST":
        update_item_form = ItemForm(request.POST,request.FILES,instance=instance)
        if update_item_form.is_valid():
            update_item_form.save()
            messages.success(request,"You have successfully updated item information.")
            return HttpResponseRedirect("/ams/items/")
    else:
        context['form'] = ItemForm(instance=instance)

    context['item_id'] = pk
    return render(request,"edit_item.html",context)

# Delete View
@login_required(login_url="/ams/")
@check_user_client_redirect
def delete_cart_view(request,pk):
    query = models.Cart.objects.filter(cart_id=pk)
    query.delete()
    messages.success(request,"You have successfully removed cart details")
    return HttpResponseRedirect("/ams/cart/")

@login_required(login_url="/ams/")
@check_user_client_redirect
def delete_transaction_view(request,pk):
    query = models.Transaction.objects.filter(id=pk)
    query.delete()
    messages.success(request,"You have successfully removed transaction details")
    return HttpResponseRedirect("/ams/transaction/")

@login_required(login_url="/ams/")
@check_user_client_redirect
def delete_supplier_view(request,pk):
    query = models.Supplier.objects.filter(supplier_id=pk)
    query.delete()
    messages.success(request,"You have successfully removed supplier details")
    return HttpResponseRedirect("/ams/supplier/")

@login_required(login_url="/ams/")
@check_user_client_redirect
def delete_trader_view(request,pk):
    query = models.CustomUser.objects.filter(id=pk)
    query.delete()
    messages.success(request,"You have successfully removed trader details")
    return HttpResponseRedirect("/ams/trader/")

@login_required(login_url="/ams/")
@check_user_client_redirect
def delete_farmer_view(request,pk):
    query = models.CustomUser.objects.filter(id=pk)
    query.delete()
    messages.success(request,"You have successfully removed farmer details")
    return HttpResponseRedirect("/ams/farmer/")

@login_required(login_url="/ams/")
@check_user_client_redirect
def delete_item_view(request,pk):
    query = models.Item.objects.filter(item_id=pk)
    query.delete()
    messages.success(request,"You have successfully removed item details")
    return HttpResponseRedirect("/ams/items/")

@login_required(login_url="/ams/")
@check_user_client_redirect
def delete_order_view(request,pk):
    query = models.Order.objects.filter(order_id=pk)
    query.delete()
    messages.success(request,"You have successfully removed order details")
    return HttpResponseRedirect("/ams/order/")